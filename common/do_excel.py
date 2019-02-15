# -*- coding:utf-8 _*-
""" 
@author:mongo
@time: 2018/12/17 
@email:3126972006@qq.com
@function： 
"""
import json
import os

import openpyxl

from common import contants


class Case:
    """
    测试用例封装类
    """

    def __init__(self):
        self.id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    file_name = None

    def __init__(self, file_name):
        try:
            # 操作的文件
            self.file_name = file_name
            # 实例化一个workbooK对象
            self.workbook = openpyxl.load_workbook(filename=file_name)
            # 异常处理如何做
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found, please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):
        sheet_name = sheet_name
        sheet = self.workbook[sheet_name]  # 获取sheet
        max_row = sheet.max_row  # 获取sheet最大行数
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.id = sheet.cell(row=r, column=1).value  # 取第r行，第1格的值
            case.title = sheet.cell(row=r, column=2).value  # 取第r行，第2格的值
            case.url = sheet.cell(row=r, column=3).value  # 取第r行，第3格的值
            case.data = sheet.cell(row=r, column=4).value  # 取第r行，第4格的值
            case.method = sheet.cell(row=r, column=5).value  # 取第r行，第5格的值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6格的值
            if type(case.expected) == int:
                case.expected = str(case.expected)
            cases.append(case)  # 将case放到cases 列表里面

        return cases  # for 循环结束后返回cases列表

    def write_result(self, sheet_name, row, actual, result):
        sheet = self.workbook[sheet_name]
        sheet.cell(row, 7).value = actual  # 写入实际结果
        sheet.cell(row, 8).value = result  # 写入执行结果，PASS or FAIL
        self.workbook.save(filename=self.file_name)


if __name__ == '__main__':
    people = '{"name":"lily","age":18,"married":false,"remark":null}'  # json格式的字符串
    print(people)
    # print(type(eval(people)))  # eval函数来尝试把上面那个字符串变成字典，不行会报错

    o_dict = json.loads(people)  # json格式的字符串序列化为Python的字典
    print(type(o_dict), o_dict)

    # 打开一个存有json格式字符串的文件
    data_json = os.path.join(contants.data_dir, "data.json")
    fp = open(data_json)
    # 使用load方法将这个文件里面的字符串转成字典
    f_dict = json.load(fp=fp)
    print(f_dict['name'])
